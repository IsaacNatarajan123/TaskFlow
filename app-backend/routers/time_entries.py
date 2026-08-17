from fastapi import APIRouter, Depends, UploadFile, File, Form
from pydantic import BaseModel
from bson import ObjectId
from database import get_collection
from auth import get_current_user
import openpyxl
import io
from datetime import datetime as dt

router = APIRouter(prefix="/time-entries", tags=["time_entries"])
time_entries_collection = get_collection("time_entries")
tasks_collection = get_collection("tasks")

class TimeEntryCreate(BaseModel):
    task_id: str
    date: str
    hours: float

@router.post("")
async def log_time(entry: TimeEntryCreate, current_user: str = Depends(get_current_user)):
    if entry.hours < 0 or entry.hours > 10:
        return {"error": "Hours must be between 0 and 10"}

    user = await get_collection("users").find_one({"_id": ObjectId(current_user)})
    if user and user.get("designation") == "CEO":
        return {"error": "CEO accounts do not participate in time logging"}

    task = await tasks_collection.find_one({"_id": ObjectId(entry.task_id)})
    if not task:
        return {"error": "Task not found"}
    if task["created_by"] != current_user:
        return {"error": "You can only log time against your own tasks"}

    client = await get_collection("clients").find_one({"_id": ObjectId(task["client_id"])})
    if client and client.get("status") == "inactive":
        return {"error": "This task's client has been deactivated. No new hours can be logged."}
    department = await get_collection("departments").find_one({"_id": ObjectId(task["department_id"])})
    if department and department.get("status") == "inactive":
        return {"error": "This task's department has been deactivated. No new hours can be logged."}

    existing = await time_entries_collection.find_one({
        "user_id": current_user, "task_id": entry.task_id, "date": entry.date
    })

    # Daily total check (hard block at 10 hours)
    day_entries = time_entries_collection.find({"user_id": current_user, "date": entry.date})
    day_total = 0.0
    async for e in day_entries:
        if existing and str(e["_id"]) == str(existing["_id"]):
            continue
        day_total += e["hours"]
    if day_total + entry.hours > 10:
        return {"error": "Daily total cannot exceed 10 hours"}

    if existing:
        await time_entries_collection.update_one({"_id": existing["_id"]}, {"$set": {"hours": entry.hours}})
        return {"message": "Entry updated"}
    else:
        result = await time_entries_collection.insert_one({
            "user_id": current_user, "task_id": entry.task_id, "date": entry.date,
            "hours": entry.hours, "submission_id": None
        })
        return {"message": "Entry logged", "id": str(result.inserted_id)}

@router.get("")
async def get_week_entries(week_start: str, current_user: str = Depends(get_current_user)):
    from datetime import datetime, timedelta
    start = datetime.strptime(week_start, "%Y-%m-%d")
    end = start + timedelta(days=6)
    cursor = time_entries_collection.find({
        "user_id": current_user,
        "date": {"$gte": start.strftime("%Y-%m-%d"), "$lte": end.strftime("%Y-%m-%d")}
    })
    entries = []
    async for e in cursor:
        e["_id"] = str(e["_id"])
        entries.append(e)
    return entries

@router.get("/team-workload")
async def team_workload(current_user: str = Depends(get_current_user)):
    from datetime import datetime, timedelta

    users_collection = get_collection("users")
    reports_cursor = users_collection.find({"manager_id": current_user})
    reports = [{"user_id": str(u["_id"]), "name": u["name"]} async for u in reports_cursor]

    if not reports:
        return {"members": [], "week_start": None, "week_end": None}

    today = datetime.now()
    week_start = (today - timedelta(days=today.weekday())).strftime("%Y-%m-%d")
    week_end = (today + timedelta(days=6 - today.weekday())).strftime("%Y-%m-%d")

    report_ids = [r["user_id"] for r in reports]
    entries_cursor = time_entries_collection.find({
        "user_id": {"$in": report_ids},
        "date": {"$gte": week_start, "$lte": week_end}
    })
    entries = [e async for e in entries_cursor]

    members = []
    for r in reports:
        member_entries = [e for e in entries if e["user_id"] == r["user_id"]]
        by_day = {}
        for e in member_entries:
            by_day[e["date"]] = by_day.get(e["date"], 0) + e["hours"]
        total = sum(by_day.values())
        members.append({
            "user_id": r["user_id"],
            "name": r["name"],
            "by_day": by_day,
            "total_hours": total,
        })

    return {"members": members, "week_start": week_start, "week_end": week_end}

@router.get("/by-submission/{submission_id}")
async def get_entries_by_submission(submission_id: str, current_user: str = Depends(get_current_user)):
    from database import get_collection as gc
    submissions = gc("weekly_submissions")
    users = gc("users")

    sub = await submissions.find_one({"_id": ObjectId(submission_id)})
    if not sub:
        return {"error": "Submission not found"}

    owner = await users.find_one({"_id": ObjectId(sub["user_id"])})
    is_owner = sub["user_id"] == current_user
    is_manager = owner and owner.get("manager_id") == current_user
    if not (is_owner or is_manager):
        return {"error": "Not authorized to view these entries"}

    cursor = time_entries_collection.find({"submission_id": submission_id})
    entries = []
    async for e in cursor:
        e["_id"] = str(e["_id"])
        entries.append(e)
    return entries

@router.post("/bulk-upload")
async def bulk_upload(file: UploadFile = File(...), week_start: str = Form(...), current_user: str = Depends(get_current_user)):
    from datetime import datetime, timedelta

    contents = await file.read()
    filename = file.filename.lower()

    if filename.endswith(".csv"):
        import csv
        decoded = contents.decode("utf-8")
        csv_reader = csv.reader(io.StringIO(decoded))
        all_rows = list(csv_reader)
    else:
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))

    header = all_rows[0]
    rows = all_rows[1:]

    # Reconstruct real dates from week_start, since the header only has day labels like "22-Wed"
    start = datetime.strptime(week_start, "%Y-%m-%d")
    week_dates = [(start + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(len(header) - 1)]

    saved = 0
    skipped = []

    for i, row in enumerate(rows, start=2):
        task_title = row[0]
        if not task_title:
            continue

        matching_tasks = await tasks_collection.count_documents({"title": task_title, "created_by": current_user})
        if matching_tasks == 0:
            skipped.append({"row": i, "reason": f"Task '{task_title}' not found"})
            continue
        if matching_tasks > 1:
            skipped.append({"row": i, "reason": f"Multiple tasks named '{task_title}' — please rename one to avoid ambiguity"})
            continue
        task = await tasks_collection.find_one({"title": task_title, "created_by": current_user})

        client = await get_collection("clients").find_one({"_id": ObjectId(task["client_id"])})
        if client and client.get("status") == "inactive":
            skipped.append({"row": i, "reason": "Task's client has been deactivated"})
            continue
        department = await get_collection("departments").find_one({"_id": ObjectId(task["department_id"])})
        if department and department.get("status") == "inactive":
            skipped.append({"row": i, "reason": "Task's department has been deactivated"})
            continue

        for col_idx, hours in enumerate(row[1:]):
            if hours is None or hours == "":
                continue
            if col_idx >= len(week_dates):
                continue
            date_str = week_dates[col_idx]

            try:
                hours_val = float(hours)
            except (ValueError, TypeError):
                skipped.append({"row": i, "reason": f"Invalid hours value for {date_str}"})
                continue

            if hours_val < 0 or hours_val > 10:
                skipped.append({"row": i, "reason": f"Hours must be between 0 and 10 ({date_str})"})
                continue

            existing = await time_entries_collection.find_one({
                "user_id": current_user, "task_id": str(task["_id"]), "date": date_str
            })
            day_entries = time_entries_collection.find({"user_id": current_user, "date": date_str})
            day_total = 0.0
            async for e in day_entries:
                if existing and str(e["_id"]) == str(existing["_id"]):
                    continue
                day_total += e["hours"]
            if day_total + hours_val > 10:
                skipped.append({"row": i, "reason": f"Would exceed 10-hour daily cap ({date_str})"})
                continue

            if existing:
                await time_entries_collection.update_one({"_id": existing["_id"]}, {"$set": {"hours": hours_val}})
            else:
                await time_entries_collection.insert_one({
                    "user_id": current_user, "task_id": str(task["_id"]), "date": date_str,
                    "hours": hours_val, "submission_id": None
                })
            saved += 1

    return {"saved": saved, "skipped": skipped}

@router.get("/download-template")
async def download_template(week_start: str, current_user: str = Depends(get_current_user)):
    from openpyxl import Workbook
    from datetime import datetime, timedelta
    import io as io_module
    from fastapi.responses import StreamingResponse

    start = datetime.strptime(week_start, "%Y-%m-%d")
    days = [(start + timedelta(days=i)) for i in range(7)]
    day_labels = [f"{d.day}-{d.strftime('%a')}" for d in days]
    day_strs = [d.strftime("%Y-%m-%d") for d in days]

    user_tasks = tasks_collection.find({"created_by": current_user, "status": {"$ne": "closed"}})
    task_list = [t async for t in user_tasks]

    existing_entries = time_entries_collection.find({
        "user_id": current_user,
        "date": {"$gte": day_strs[0], "$lte": day_strs[-1]}
    })
    entry_map = {}
    async for e in existing_entries:
        entry_map[(e["task_id"], e["date"])] = e["hours"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Log Time"
    ws.append(["Task"] + day_labels)

    for t in task_list:
        row = [t["title"]]
        for ds in day_strs:
            row.append(entry_map.get((str(t["_id"]), ds), ""))
        ws.append(row)

    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max_length + 3

    output = io_module.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=timesheet_template.xlsx"}
    )

@router.delete("/{task_id}/{date}")
async def delete_time_entry(task_id: str, date: str, current_user: str = Depends(get_current_user)):
    result = await time_entries_collection.delete_one({
        "user_id": current_user, "task_id": task_id, "date": date
    })
    if result.deleted_count == 0:
        return {"error": "Entry not found"}
    return {"message": "Entry deleted"}